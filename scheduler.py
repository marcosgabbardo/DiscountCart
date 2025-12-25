#!/usr/bin/env python3
"""
Scheduler para atualização automática de preços.
Executa às 8:00 da manhã e gera relatório Excel.
Suporta Zaffari e Carrefour.
Alertas baseados em desvio padrão (1DP e 2DP) para períodos de 30, 90 e 180 dias.
"""

import schedule
import time
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from services import ProductService, AlertService
from database.models import AlertType, Store


def generate_excel_report(
    products_with_changes: List[Dict[str, Any]],
    std_dev_alerts: Dict[str, Dict[str, List[Dict]]]
) -> str:
    """
    Gera relatório Excel com produtos que tiveram alteração de preço.
    Destaca ofertas baseadas em desvio padrão.
    Retorna o caminho do arquivo gerado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Preços"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    std_2_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde - 2 DP
    std_1_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo - 1 DP
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "ID", "Loja", "Produto", "Preço Anterior", "Preço Atual", "Variação",
        "Média 30d", "DP 30d", "Média 90d", "DP 90d", "Média 180d", "DP 180d", "Alerta"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Collect product IDs that are in 2 std dev (best deals) and 1 std dev
    products_2_std = set()
    products_1_std = set()

    for period in ['30d', '90d', '180d']:
        for item in std_dev_alerts['2_std_dev'].get(period, []):
            products_2_std.add(item['product'].id)
        for item in std_dev_alerts['1_std_dev'].get(period, []):
            products_1_std.add(item['product'].id)

    # Dados
    row = 2
    for item in products_with_changes:
        product = item['product']
        prev_price = item.get('previous_price')
        variation = item.get('variation', 0)

        # Get stats for different periods
        stats_30 = item.get('stats_30', {})
        stats_90 = item.get('stats_90', {})
        stats_180 = item.get('stats_180', {})

        # Determinar tipo de alerta
        alert_type = ""
        row_fill = None

        if product.id in products_2_std:
            alert_type = "🔥 2 DESVIOS PADRÃO"
            row_fill = std_2_fill
        elif product.id in products_1_std:
            alert_type = "✅ 1 DESVIO PADRÃO"
            row_fill = std_1_fill

        # Get store display name
        store_name = product.store.display_name if hasattr(product.store, 'display_name') else str(product.store)

        # Preencher linha
        ws.cell(row=row, column=1, value=product.id).border = border
        ws.cell(row=row, column=2, value=store_name).border = border
        ws.cell(row=row, column=3, value=product.title[:45] if product.title else "").border = border
        ws.cell(row=row, column=4, value=float(prev_price) if prev_price else None).border = border
        ws.cell(row=row, column=5, value=float(product.current_price) if product.current_price else None).border = border
        ws.cell(row=row, column=6, value=f"{variation:+.2f}%").border = border
        ws.cell(row=row, column=7, value=float(stats_30.get('avg_price', 0)) if stats_30.get('avg_price') else None).border = border
        ws.cell(row=row, column=8, value=float(stats_30.get('std_deviation', 0)) if stats_30.get('std_deviation') else None).border = border
        ws.cell(row=row, column=9, value=float(stats_90.get('avg_price', 0)) if stats_90.get('avg_price') else None).border = border
        ws.cell(row=row, column=10, value=float(stats_90.get('std_deviation', 0)) if stats_90.get('std_deviation') else None).border = border
        ws.cell(row=row, column=11, value=float(stats_180.get('avg_price', 0)) if stats_180.get('avg_price') else None).border = border
        ws.cell(row=row, column=12, value=float(stats_180.get('std_deviation', 0)) if stats_180.get('std_deviation') else None).border = border
        ws.cell(row=row, column=13, value=alert_type).border = border

        # Aplicar cor se tem alerta
        if row_fill:
            for col in range(1, 14):
                ws.cell(row=row, column=col).fill = row_fill

        row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 25

    # Salvar arquivo
    filename = f"relatorio_precos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    return filename


def run_daily_update():
    """Executa atualização diária e gera relatório."""
    print(f"\n{'='*60}")
    print(f"⏰ Atualização diária iniciada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print('='*60)

    product_service = ProductService()
    alert_service = AlertService()

    # Guardar preços anteriores
    products = product_service.get_all_products(active_only=True)
    previous_prices = {p.id: p.current_price for p in products}

    # Atualizar preços
    print("\n📦 Atualizando preços...")
    updated = product_service.update_all_prices()
    print(f"✅ {len(updated)} produto(s) atualizado(s)")

    # Identificar produtos com alteração
    products_with_changes = []
    for product in updated:
        prev_price = previous_prices.get(product.id)
        if prev_price and product.current_price and prev_price != product.current_price:
            variation = float((product.current_price - prev_price) / prev_price * 100)

            # Get stats for all periods
            stats_30 = product_service.get_std_deviation(product.id, 30)
            stats_90 = product_service.get_std_deviation(product.id, 90)
            stats_180 = product_service.get_std_deviation(product.id, 180)

            products_with_changes.append({
                'product': product,
                'previous_price': prev_price,
                'variation': variation,
                'stats_30': {'avg_price': stats_30[0], 'std_deviation': stats_30[1]} if stats_30 else {},
                'stats_90': {'avg_price': stats_90[0], 'std_deviation': stats_90[1]} if stats_90 else {},
                'stats_180': {'avg_price': stats_180[0], 'std_deviation': stats_180[1]} if stats_180 else {},
            })

    # Obter alertas de desvio padrão
    std_dev_alerts = alert_service.check_std_deviation_alerts()

    # Gerar relatório Excel se houver alterações
    if products_with_changes:
        print(f"\n📊 {len(products_with_changes)} produto(s) com alteração de preço")
        filename = generate_excel_report(products_with_changes, std_dev_alerts)
        print(f"📄 Relatório gerado: {filename}")
    else:
        print("\n📊 Nenhuma alteração de preço detectada")

    # Resumo de alertas de desvio padrão
    print("\n" + "=" * 60)
    print("📊 ALERTAS DE DESVIO PADRÃO")
    print("=" * 60)

    # 2 desvios padrão (ofertas excepcionais)
    total_2_std = 0
    for period in ['30d', '90d', '180d']:
        items = std_dev_alerts['2_std_dev'].get(period, [])
        total_2_std += len(items)
        if items:
            print(f"\n🔥 2 DP ({period}): {len(items)} produto(s)")
            for item in items[:5]:  # Mostrar apenas os 5 primeiros
                p = item['product']
                store_name = p.store.display_name if hasattr(p.store, 'display_name') else str(p.store)
                print(f"   • [{store_name}] {p.title[:35]}: R$ {p.current_price:.2f}")

    # 1 desvio padrão (boas ofertas)
    total_1_std = 0
    for period in ['30d', '90d', '180d']:
        items = std_dev_alerts['1_std_dev'].get(period, [])
        total_1_std += len(items)
        if items:
            print(f"\n✅ 1 DP ({period}): {len(items)} produto(s)")
            for item in items[:5]:  # Mostrar apenas os 5 primeiros
                p = item['product']
                store_name = p.store.display_name if hasattr(p.store, 'display_name') else str(p.store)
                print(f"   • [{store_name}] {p.title[:35]}: R$ {p.current_price:.2f}")

    print(f"\n📈 Resumo: {total_2_std} ofertas excepcionais (2DP) | {total_1_std} boas ofertas (1DP)")

    print(f"\n{'='*60}")
    print(f"✅ Atualização concluída: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print('='*60)


def main():
    """Inicia o scheduler."""
    print("🕐 Scheduler iniciado")
    print("   Atualização programada para 08:00 todos os dias")
    print("   Alertas baseados em desvio padrão:")
    print("   - 1 DP: Boas ofertas")
    print("   - 2 DP: Ofertas excepcionais")
    print("   - Períodos: 30, 90 e 180 dias")
    print("   Pressione Ctrl+C para parar\n")

    # Agendar para 8:00
    schedule.every().day.at("08:00").do(run_daily_update)

    # Loop principal
    while True:
        schedule.run_pending()
        time.sleep(60)  # Verifica a cada minuto


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == '--now':
        # Executar imediatamente (para testes)
        run_daily_update()
    else:
        main()
